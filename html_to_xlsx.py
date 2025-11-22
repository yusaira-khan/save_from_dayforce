import os.path
from strip_html import read_soup
from openpyxl import Workbook


def clear(d):
    for f in os.listdir(d):
        os.remove(os.path.join(d, f))


def main():
    input_dir = "2025-html"
    wb = Workbook()
    for f in os.listdir(input_dir):
        soup = read_soup(os.path.join(input_dir, f))
        output_sheet_name =  f.replace(".html", "" )
        ws = wb.create_sheet(output_sheet_name)
        write_tables(soup, ws)
    wb.save("qcom.xlsx")


TABLE_SEPARATION = 3
def write_tables(soup, ws):
    tables = soup.find_all("table")
    current_row = 1
    for table in tables:
        for html_row in table.find_all("tr"):
            current_col = 1
            for item in html_row:
                txt = item.get_text().strip()
                ws.cell(current_row, current_col, txt)
                current_col += 1
            current_row+=1

if __name__ == "__main__":
    main()